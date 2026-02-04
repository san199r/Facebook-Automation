import os
import re
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException

from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
START_URL = "https://www.facebook.com/UseApolloIo/followers/"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

EXCEL_FILE = os.path.join(
    OUTPUT_DIR,
    "apollo_page_fb_followers.xlsx"
)

HEADERS = [
    "S.No",
    "Facebook Name",
    "Facebook Page URL",
    "Location",
    "Phone",
    "Email",
    "Website",
    "External Facebook",
    "External LinkedIn",
    "External Instagram",
]


# ================= SAFE PRINT =================
def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="ignore").decode())


# ================= DRIVER =================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.set_page_load_timeout(60)
    return driver


# ================= COOKIES =================
def load_facebook_cookies(driver, cookie_file):
    driver.get("https://www.facebook.com/")
    time.sleep(3)

    with open(cookie_file, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue

            domain, flag, path, secure, expiry, name, value = line.strip().split("\t")

            if not domain.endswith("facebook.com"):
                continue

            cookie = {
                "name": name,
                "value": value,
                "domain": domain,
                "path": path,
            }

            if expiry.isdigit():
                cookie["expiry"] = int(expiry)

            driver.add_cookie(cookie)

    driver.refresh()
    time.sleep(5)


# ================= EXCEL (RESUME) =================
def init_or_resume_excel():
    collected_urls = set()

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        start_sno = ws.max_row

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[2]:
                collected_urls.add(row[2])

        safe_print(f"Resuming Excel with {len(collected_urls)} existing records")

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Apollo Followers"

        bold = Font(bold=True)
        for c, h in enumerate(HEADERS, start=1):
            cell = ws.cell(1, c, h)
            cell.font = bold

        start_sno = 1
        safe_print("Creating new Excel file")

    return wb, ws, collected_urls, start_sno


# ================= HELPERS =================
def normalize(text):
    return re.sub(r"\s+", " ", (text or "").strip())


def is_valid_name(name):
    bad = {
        "followers", "following", "about", "mentions", "reviews",
        "reels", "photos", "home", "friends", "messages"
    }
    return name and name.lower() not in bad and len(name) > 2


# ================= MAIN =================
def scrape_followers():
    driver = init_driver()
    wait = WebDriverWait(driver, 30)

    safe_print("Loading Facebook cookies")
    load_facebook_cookies(driver, COOKIE_FILE)

    if "login" in driver.current_url.lower():
        raise Exception("Cookie login failed")

    safe_print("Login successful")

    driver.get(START_URL)
    time.sleep(5)

    wb, ws, collected, sno = init_or_resume_excel()

    no_new_rounds = 0
    MAX_NO_NEW = 15

    safe_print("Collecting Apollo followers")

    while no_new_rounds < MAX_NO_NEW:
        found_this_round = 0

        anchors = driver.find_elements(
            By.XPATH,
            "//div[@role='main']//a[contains(@href,'/profile.php') or contains(@href,'/people/')]"
        )

        for a in anchors:
            try:
                name = normalize(a.text)
                href = a.get_attribute("href")

                if not is_valid_name(name):
                    continue

                if href in collected:
                    continue

                collected.add(href)

                ws.append([
                    sno,
                    name,
                    href,
                    "",  # Location
                    "",  # Phone
                    "",  # Email
                    "",  # Website
                    "",  # External Facebook
                    "",  # External LinkedIn
                    "",  # External Instagram
                ])

                safe_print(f"Collected {sno}: {name}")

                sno += 1
                found_this_round += 1

            except StaleElementReferenceException:
                continue

        if found_this_round == 0:
            no_new_rounds += 1
            safe_print(f"No new followers ({no_new_rounds}/{MAX_NO_NEW})")
        else:
            no_new_rounds = 0

        driver.execute_script("window.scrollBy(0, 1600);")
        time.sleep(2)

    wb.save(EXCEL_FILE)
    safe_print(f"Excel saved at: {EXCEL_FILE}")

    driver.quit()
    safe_print("Browser closed")


if __name__ == "__main__":
    scrape_followers()
