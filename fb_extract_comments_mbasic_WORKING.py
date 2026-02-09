import os
import time
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
SOURCE = "FB"
KEYWORD = "PROBATE"

INPUT_EXCEL = "output/fb_probate_ALL_posts_20260209_110948.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
SCREENSHOT_DIR = os.path.join(OUTPUT_DIR, "screenshots")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR, f"fb_{KEYWORD}_COMMENTS_MBASIC_{TIMESTAMP}.xlsx"
)


# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-notifications")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )


# ================= LOAD COOKIES =================
def load_cookies(driver):
    driver.get("https://mbasic.facebook.com/")
    time.sleep(4)

    if not os.path.exists(COOKIE_FILE):
        print("Cookie file not found")
        return

    with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if line.strip() and not line.startswith("#"):
                parts = line.strip().split("\t")
                if len(parts) >= 7:
                    driver.add_cookie({
                        "name": parts[5],
                        "value": parts[6],
                        "domain": ".facebook.com"
                    })

    driver.refresh()
    time.sleep(6)
    print("Cookies loaded (mbasic)")


def to_mbasic(url):
    return url.replace("www.facebook.com", "mbasic.facebook.com")


# ================= EXPAND COMMENTS =================
def expand_all_comments(driver):
    while True:
        links = driver.find_elements(
            By.XPATH, "//a[contains(text(),'View more comments')]"
        )
        if not links:
            break
        try:
            links[0].click()
            time.sleep(3)
        except:
            break


# ================= MAIN =================
def run():
    driver = init_driver()
    load_cookies(driver)

    wb_in = load_workbook(INPUT_EXCEL)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Comments"

    headers = [
        "Source",
        "Keyword",
        "Commentator Name",
        "Url to find the Comment",
        "Comment"
    ]
    ws_out.append(headers)
    for cell in ws_out[1]:
        cell.font = Font(bold=True)

    try:
        for idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), 1):
            post_url = row[1]
            mbasic_url = to_mbasic(post_url)

            print(f"[{idx}] Opening: {mbasic_url}")
            driver.get(mbasic_url)
            time.sleep(5)

            driver.save_screenshot(
                os.path.join(SCREENSHOT_DIR, f"post_{idx:02d}_open.png")
            )

            expand_all_comments(driver)

            driver.save_screenshot(
                os.path.join(SCREENSHOT_DIR, f"post_{idx:02d}_expanded.png")
            )

            comment_blocks = driver.find_elements(
                By.XPATH, "//div[starts-with(@id,'comment_')]"
            )

            extracted_any = False

            for c in comment_blocks:
                try:
                    name = c.find_element(By.XPATH, ".//strong").text.strip()
                except:
                    continue

                spans = c.find_elements(By.XPATH, ".//span")
                comment_text = " ".join(
                    s.text.strip() for s in spans if s.text.strip()
                )

                if comment_text:
                    ws_out.append([
                        SOURCE,
                        KEYWORD,
                        name,
                        post_url,
                        comment_text
                    ])
                    extracted_any = True

            if not extracted_any:
                ws_out.append([
                    SOURCE,
                    KEYWORD,
                    "NO_COMMENTS",
                    post_url,
                    "NO_COMMENTS"
                ])

    finally:
        wb_out.save(OUTPUT_EXCEL)
        driver.quit()

        print("===================================")
        print("DONE")
        print(f"Excel saved: {OUTPUT_EXCEL}")
        print(f"Screenshots saved: {SCREENSHOT_DIR}")
        print("===================================")


if __name__ == "__main__":
    run()
