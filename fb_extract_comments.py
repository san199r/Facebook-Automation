import os
import time
from glob import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
OUTPUT_DIR = "output"
FINAL_EXCEL = "fb_final_output.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

MAX_POSTS = 20
SCROLL_COMMENTS = 6

SOURCE = "Facebook"
KEYWORD = "probate"

RESPONDER_FIRST_NAME = "Santosh"
RESPONDER_LAST_NAME = "P"


# ================= SAFE PRINT =================
def safe_print(text):
    try:
        print(text)
    except Exception:
        print(text.encode("ascii", errors="ignore").decode())


# ================= SIMPLE CHATGPT-LIKE LOGIC =================
def generate_comment_response(comment):
    text = comment.lower()

    if any(w in text for w in ["help", "need", "advice", "guidance"]):
        return (
            "Thanks for reaching out. This is a common situation and there are clear options "
            "available. Feel free to share a bit more detail and I’ll be happy to help."
        )

    if any(w in text for w in ["interested", "dm", "contact"]):
        return (
            "Thanks for your interest. I’ve shared some helpful information. "
            "Please feel free to reach out if you have any questions."
        )

    if any(w in text for w in ["thank", "thanks"]):
        return "You’re welcome. Let me know if you need any further information."

    return (
        "Thanks for your comment. If you’d like more information or guidance, "
        "feel free to ask."
    )


def generate_responder_reply(comment):
    return generate_comment_response(comment) + " Looking forward to connecting."


# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )


# ================= LOAD COOKIES =================
def load_driver_with_cookies():
    driver = init_driver()
    driver.get("https://www.facebook.com/")
    time.sleep(4)

    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    parts = line.strip().split("\t")
                    if len(parts) >= 7:
                        driver.add_cookie({
                            "name": parts[5],
                            "value": parts[6],
                            "domain": parts[0]
                        })
        driver.refresh()
        time.sleep(6)
        safe_print("Cookies loaded successfully")
    else:
        safe_print("Cookie file not found")

    return driver


# ================= FIND POST-URL EXCEL =================
def get_latest_post_excel():
    files = glob(os.path.join(OUTPUT_DIR, "*.xlsx"))
    for f in sorted(files, key=os.path.getmtime, reverse=True):
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        if "Post URL" in headers:
            safe_print(f"Using input Excel: {f}")
            return f
    raise Exception("No Excel with 'Post URL' column found")


# ================= READ POST URLS =================
def read_post_urls():
    excel = get_latest_post_excel()
    wb = load_workbook(excel)
    ws = wb.active

    post_col = [c.value for c in ws[1]].index("Post URL")

    urls = []
    for r in ws.iter_rows(min_row=2):
        if len(urls) >= MAX_POSTS:
            break
        if r[post_col].value:
            urls.append(r[post_col].value)

    safe_print(f"Loaded {len(urls)} post URLs")
    return urls


# ================= EXTRACT COMMENTS =================
def extract_comments(driver, post_url, ws):
    safe_print(f"Opening post: {post_url}")
    driver.get(post_url)
    time.sleep(8)

    for _ in range(SCROLL_COMMENTS):
        driver.execute_script("window.scrollBy(0,1500);")
        time.sleep(3)

    comments = driver.find_elements(
        By.XPATH,
        "//div[@role='article']//span[contains(@class,'x1lliihq')]"
    )

    for c in comments:
        try:
            text = c.text.strip()
            if not text:
                continue

            parent = c.find_element(By.XPATH, "./ancestor::div[@role='article']")
            profile = parent.find_element(By.XPATH, ".//a[@role='link']")

            full_name = profile.text.strip()
            parts = full_name.split(" ", 1)
            first = parts[0]
            last = parts[1] if len(parts) > 1 else ""

            ai_comment = generate_comment_response(text)
            ai_reply = generate_responder_reply(text)

            ws.append([
                SOURCE,
                KEYWORD,
                "",
                first,
                last,
                post_url,
                text,
                ai_comment,
                RESPONDER_FIRST_NAME,
                RESPONDER_LAST_NAME,
                ai_reply,
                ai_reply,
            ])

        except Exception:
            continue


# ================= MAIN =================
def run():
    driver = load_driver_with_cookies()
    post_urls = read_post_urls()

    wb = Workbook()
    ws = wb.active
    ws.title = "Final Data"

    headers = [
        "Source",
        "Keyword",
        "Group",
        "First Name of the commenter",
        "Last Name of the commenter",
        "Url to find the Comment",
        "Comment",
        "CHATGPT response for Comment",
        "First Name of the Responder1",
        "Last Name of the Responder1",
        "Comment",
        "CHATGPT response for Comments Reponse",
    ]

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for url in post_urls:
        extract_comments(driver, url, ws)

    wb.save(FINAL_EXCEL)
    safe_print(f"FINAL EXCEL CREATED: {FINAL_EXCEL}")
    driver.quit()


if __name__ == "__main__":
    run()
