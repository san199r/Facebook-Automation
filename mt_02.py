import os
import time
import sys
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ================= CONFIG =================
INPUT_EXCEL = "clean_posts.xlsx"
COOKIE_FILE = "cookies/facebook_cookies.txt"

# ================= FORCE UTF-8 OUTPUT =================
try:
    sys.stdout.reconfigure(encoding='utf-8')
except:
    pass

# ================= SAFE PRINT =================
def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="ignore").decode())

# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

# ================= LOAD COOKIES =================
def load_cookies(driver):
    driver.get("https://www.facebook.com/")
    time.sleep(5)

    if not os.path.exists(COOKIE_FILE):
        print("Cookie file not found!")
        return

    with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as file:
        for line in file:
            if line.startswith("#") or not line.strip():
                continue

            parts = line.strip().split("\t")
            if len(parts) >= 7:
                cookie = {
                    "domain": parts[0],
                    "name": parts[5],
                    "value": parts[6],
                    "path": parts[2],
                }
                try:
                    driver.add_cookie(cookie)
                except:
                    pass

    driver.refresh()
    time.sleep(5)

    print("Login URL:", driver.current_url)
    print("Page Title:", driver.title)

# ================= READ POST URLS =================
def read_urls():
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            urls.append(row[0])

    print("Total Posts:", len(urls))
    return urls

# ================= EXPAND COMMENTS =================
def expand_comments(driver):
    while True:
        try:
            more = driver.find_element(By.XPATH, "//span[contains(text(),'View more comments')]")
            driver.execute_script("arguments[0].click();", more)
            time.sleep(2)
        except:
            break

# ================= EXTRACT COMMENTS =================
def extract_comments(driver, post_url):

    print("\n========================================")
    print("Opening Post:", post_url)
    print("========================================")

    driver.get(post_url)
    time.sleep(6)

    # Scroll down multiple times
    for _ in range(5):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)

    expand_comments(driver)

    # Extract comment blocks via JS
    comments = driver.execute_script("""
        let results = [];
        document.querySelectorAll('div[dir="auto"]').forEach(el => {
            if(el.innerText.length > 5) {
                results.push(el.innerText);
            }
        });
        return results;
    """)

    print("Total detected blocks:", len(comments))

    for i, c in enumerate(comments, 1):
        print(f"\n--- Comment {i} ---")
        safe_print(c)

# ================= MAIN =================
def run():
    driver = init_driver()

    # Login using cookies
    load_cookies(driver)

    # Read URLs
    urls = read_urls()

    # Loop through posts
    for i, url in enumerate(urls, 1):
        print(f"\nProcessing {i}/{len(urls)}")
        extract_comments(driver, url)
        time.sleep(5)

    print("\nAll posts processed successfully.")
    driver.quit()

if __name__ == "__main__":
    run()
