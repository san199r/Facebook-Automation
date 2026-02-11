import re
from openpyxl import load_workbook, Workbook

INPUT_FILE = "facebook_fixed.xlsx"   # your messy extension file
OUTPUT_FILE = "clean_posts.xlsx"

def extract_clean_urls():
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    clean_urls = set()

    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):

                # Match photo post URLs
                match = re.search(r"https://www\.facebook\.com/photo/\?fbid=\d+", cell)
                if match:
                    clean_urls.add(match.group(0))

                # Match permalink style URLs
                match2 = re.search(r"https://www\.facebook\.com/.+?/posts/\d+", cell)
                if match2:
                    clean_urls.add(match2.group(0))

    return clean_urls

def save_clean_file(urls):
    wb = Workbook()
    ws = wb.active
    ws.append(["Post URL"])

    for url in urls:
        ws.append([url])

    wb.save(OUTPUT_FILE)

if __name__ == "__main__":
    urls = extract_clean_urls()
    save_clean_file(urls)

    print("Total Clean URLs Found:", len(urls))
    print("Saved as:", OUTPUT_FILE)
