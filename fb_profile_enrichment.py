import os
import time
import re
from datetime import datetime
from urllib.parse import unquote, urlparse, parse_qs

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException
)
from webdriver_manager.chrome import ChromeDriverManager


# ================= SAFE PRINT =================
def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="ignore").decode())


def clean_text(text):
    if not text:
        return ""
    return str(text).strip()


# ================= CONFIG =================
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR,
    f"facebook_profiles_enriched_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

HEADERS = [
    "Name",
    "Profile URL",
    "Address",
    "Email",
    "Phone",
    "YouTube",
    "Instagram",
    "Website",
    "LinkedIn",
    "Twitter"
]


# ================= INPUT EXCEL =================
def get_latest_followers_excel():
    files = [
        f for f in os.listdir(OUTPUT_DIR)
        if f.startswith("facebook_followers_") and f.endswith(".xlsx")
    ]
    if not files:
        raise FileNotFoundError("No facebook_followers_*.xlsx found")

    files.sort(reverse=True)
    latest = os.path.join(OUTPUT_DIR, files[0])
    safe_print(f"Using input Excel: {latest}")
    return latest


INPUT_EXCEL = get_latest_followers_excel()


# ================= RETRY =================
def retry(action, retries=3, wait=1):
    for _ in range(retries):
        try:
            return action()
        except StaleElementReferenceException:
            time.sleep(wait)
    return None


# ================= DRIVER =================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.set_page_load_timeout(60)
    return driver


# ================= COOKIES =================
def load_facebook_cookies(driver):
    driver.get("https://www.facebook.com/")
    time.sleep(3)

    with open(COOKIE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue

            domain, flag, path, secure, expiry, name, value = line.strip().split("\t")

            cookie = {
                "name": name,
                "value": value,
                "domain": domain,
                "path": path
            }
            if expiry.isdigit():
                cookie["expiry"] = int(expiry)

            driver.add_cookie(cookie)

    driver.refresh()
    time.sleep(5)


# ================= OUTPUT EXCEL =================
def init_output_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"

    bold = Font(bold=True)
    for col, h in enumerate(HEADERS, start=1):
        ws.cell(1, col, h).font = bold

    return wb, ws


# ================= ABOUT =================
def open_about_and_scroll(driver):
    try:
        retry(lambda: driver.find_element(
            By.XPATH,
            "//a[contains(@href,'/about') or .//span[normalize-space()='About']]"
        ).click())
        time.sleep(3)
    except Exception:
        driver.get(driver.current_url.rstrip("/") + "/about")
        time.sleep(3)

    for _ in range(5):
        driver.execute_script("window.scrollBy(0,700)")
        time.sleep(1.5)


# ================= CONTACT =================
def extract_contact_info(driver):
    data = {"email": "", "phone": "", "address": ""}

    try:
        blocks = driver.find_elements(
            By.XPATH,
            "//div[@role='main']//span | //div[@role='main']//div"
        )
    except StaleElementReferenceException:
        return data

    for b in blocks:
        try:
            raw = b.text.strip()
        except StaleElementReferenceException:
            continue

        for t in raw.split("\n"):
            tl = t.lower()

            if not data["email"]:
                m = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", t)
                if m:
                    data["email"] = m.group(0)

            if not data["phone"]:
                m = re.search(r"\+?\d[\d\s\-]{7,}", t)
                if m:
                    data["phone"] = m.group(0)

            if (
                not data["address"]
                and len(t) > 10
                and "@" not in t
                and not t.startswith("http")
            ):
                if any(x in tl for x in ["street", "road", "city", "state", "india", "bangladesh"]):
                    data["address"] = t.strip()

    return data


# ================= LINKS (MENTIONS REMOVED) =================
def extract_links(driver):
    links = {
        "youtube": "",
        "instagram": "",
        "website": "",
        "linkedin": "",
        "twitter": ""
    }

    try:
        anchors = driver.find_elements(By.XPATH, "//div[@role='main']//a[@href]")
    except StaleElementReferenceException:
        return links

    for a in anchors:
        try:
            href = a.get_attribute("href")
        except StaleElementReferenceException:
            continue

        if not href or "/mentions" in href.lower():
            continue

        final_url = href

        if "l.facebook.com/l.php" in href:
            qs = parse_qs(urlparse(href).query)
            if "u" in qs:
                final_url = unquote(qs["u"][0])

        h = final_url.lower()

        if "youtube.com" in h and not links["youtube"]:
            links["youtube"] = final_url
        elif "instagram.com" in h and not links["instagram"]:
            links["instagram"] = final_url
        elif "linkedin.com" in h and not links["linkedin"]:
            links["linkedin"] = final_url
        elif ("twitter.com" in h or "x.com" in h) and not links["twitter"]:
            links["twitter"] = final_url
        elif (
            not links["website"]
            and "facebook.com" not in h
            and not any(s in h for s in ["youtube", "instagram", "linkedin", "twitter", "x.com"])
        ):
            links["website"] = final_url

    return links


# ================= MAIN =================
def enrich_profiles():
    driver = init_driver()
    load_facebook_cookies(driver)

    in_wb = load_workbook(INPUT_EXCEL)
    in_ws = in_wb.active

    out_wb, out_ws = init_output_excel()

    for row in range(2, in_ws.max_row + 1):
        name = clean_text(in_ws.cell(row, 2).value)
        profile_url = in_ws.cell(row, 3).value

        if not profile_url or "/mentions" in profile_url.lower():
            safe_print(f"[SKIPPED MENTIONS] {profile_url}")
            continue

        safe_print(f"[ROW {row}] {name} -> {profile_url}")

        try:
            driver.get(profile_url)
            time.sleep(3)

            open_about_and_scroll(driver)

            contact = extract_contact_info(driver)
            links = extract_links(driver)

            out_ws.append([
                name,
                profile_url,
                contact["address"],
                contact["email"],
                contact["phone"],
                links["youtube"],
                links["instagram"],
                links["website"],
                links["linkedin"],
                links["twitter"]
            ])

        except Exception as e:
            safe_print(f"[SKIPPED ROW {row}] {e}")

    out_wb.save(OUTPUT_EXCEL)
    safe_print(f"✅ Output saved: {OUTPUT_EXCEL}")

    driver.quit()


# ================= RUN =================
if __name__ == "__main__":
    enrich_profiles()
