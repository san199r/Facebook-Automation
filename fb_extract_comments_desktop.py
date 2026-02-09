import os
import time
import re
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
SOURCE = "FB"
KEYWORD = "PROBATE"

INPUT_EXCEL = "output/fb_probate_ALL_posts_20260209_200346.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
SCREENSHOT_DIR = os.path.join(OUTPUT_DIR, "screenshots")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR, f"fb_BATCH_comments_FINAL_{TIMESTAMP}.xlsx"
)


# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-notifications")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )


# ================= LOAD COOKIES =================
def load_cookies(driver):
    driver.get("https://mbasic.facebook.com/")
    time.sleep(4)

    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    parts = line.strip().split("\t")
                    if len(parts) >= 7:
                        driver.add_cookie({
                            "name": parts[5],
                            "value": parts[6],
                            "domain": ".facebook.com"
                        })

    driver.refresh()
    time.sleep(5)


def to_mbasic(url):
    return url.replace("www.facebook.com", "mbasic.facebook.com")


# ================= LOAD ALL COMMENTS =================
def load_all_comments(driver, idx):
    for _ in range(30):
        clicked = False
        links = driver.find_elements(By.XPATH, "//a[contains(text(),'View more comments')]")
        for l in links:
            try:
                l.click()
                clicked = True
                time.sleep(2)
            except:
                pass

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(2)

        if not clicked:
            break

    driver.save_screenshot(
        os.path.join(SCREENSHOT_DIR, f"post_{idx:03d}_final.png")
    )


# ================= EXPAND REPLIES =================
def expand_all_replies(driver):
    for _ in range(20):
        reply_links = driver.find_elements(By.XPATH, "//a[contains(text(),'reply')]")
        if not reply_links:
            break
        for link in reply_links:
            try:
                link.click()
                time.sleep(1.5)
            except:
                pass


# ================= FILTER RULES =================
IGNORE_PHRASES = [
    "like",
    "reply",
    "share",
    "comment",
    "most relevant",
    "comments",
    "write a comment",
    "view more comments",
    "view all",
    "new notification",
    "this photo is from a post",
    "no comments yet",
    "all reactions",
    "replies",
    "edited",
]

def is_valid_name(text):
    if not text:
        return False
    if len(text.split()) < 2 or len(text.split()) > 4:
        return False
    if not re.match(r"^[A-Za-z .'-]+$", text):
        return False
    bad_start = (
        "view", "new", "all", "no", "this", "part",
        "1d", "2d", "3d"
    )
    return not text.lower().startswith(bad_start)

def is_valid_comment(text):
    if not text or len(text.split()) < 3:
        return False
    low = text.lower()
    if any(p in low for p in IGNORE_PHRASES):
        return False
    return True


# ================= PARSE BODY TEXT =================
def parse_comments_from_body(body_text):
    lines = [l.strip() for l in body_text.splitlines() if l.strip()]

    cleaned = []
    for line in lines:
        low = line.lower()
        if any(p in low for p in IGNORE_PHRASES):
            continue
        if re.match(r"\d+w", low):
            continue
        if re.match(r"\d+\s+of\s+\d+", low):
            continue
        if line in {"·", ".", "…"}:
            continue
        if line.isdigit():
            continue
        cleaned.append(line)

    results = []
    last_parent = ""

    i = 0
    while i < len(cleaned) - 1:
        name = cleaned[i]
        comment = cleaned[i + 1]

        if not is_valid_name(name) or not is_valid_comment(comment):
            i += 1
            continue

        if "replied" in name.lower():
            commenter = name.replace("replied", "").strip()
            results.append((commenter, comment, "REPLY", last_parent))
        else:
            last_parent = name
            results.append((name, comment, "COMMENT", ""))

        i += 2

    return results


# ================= MAIN =================
def run():
    wb_in = load_workbook(INPUT_EXCEL)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Comments"

    headers = [
        "Source",
        "Keyword",
        "Commentator",
        "Comment",
        "Comment Type",
        "Parent Commentator",
        "Post URL"
    ]
    ws_out.append(headers)
    for c in ws_out[1]:
        c.font = Font(bold=True)

    driver = init_driver()
    load_cookies(driver)

    try:
        for idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), 1):
            post_url = row[1]
            if not post_url:
                continue

            mbasic_url = to_mbasic(post_url)
            print(f"[{idx}] Processing:", mbasic_url)

            driver.get(mbasic_url)
            time.sleep(5)

            load_all_comments(driver, idx)
            expand_all_replies(driver)
            load_all_comments(driver, idx)

            body_text = driver.find_element("tag name", "body").text
            parsed = parse_comments_from_body(body_text)

            for name, comment, ctype, parent in parsed:
                ws_out.append([
                    SOURCE,
                    KEYWORD,
                    name,
                    comment,
                    ctype,
                    parent,
                    post_url
                ])

    finally:
        wb_out.save(OUTPUT_EXCEL)
        driver.quit()

        print("===================================")
        print("DONE – CLEAN COMMENTS ONLY")
        print("Excel:", OUTPUT_EXCEL)
        print("Screenshots:", SCREENSHOT_DIR)
        print("===================================")


if __name__ == "__main__":
    run()
