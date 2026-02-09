import os
import time
import re
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
SOURCE = "FB"
KEYWORD = "PROBATE"

INPUT_EXCEL = "output/fb_probate_ALL_posts_20260209_110948.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, f"fb_CLEAN_comments_{TIMESTAMP}.xlsx")


# ================= DRIVER & COOKIES =================
def init_driver():
    options = Options()
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-notifications")
    # Using a mobile user-agent makes the mbasic site more stable
    options.add_argument("user-agent=Mozilla/5.0 (Linux; Android 10; SM-G960U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.181 Mobile Safari/537.36")
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

def load_cookies(driver):
    driver.get("https://mbasic.facebook.com/")
    time.sleep(4)
    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    parts = line.strip().split("\t")
                    if len(parts) >= 7:
                        driver.add_cookie({
                            "name": parts[5],
                            "value": parts[6],
                            "domain": ".facebook.com"
                        })
    driver.refresh()
    time.sleep(3)


# ================= CONTENT EXPANSION =================
def expand_everything(driver):
    """Clicks 'View more' and expands 'See more' text for full stories."""
    # 1. Click 'View more comments' up to 20 times
    for _ in range(20):
        try:
            # Look for links containing 'pager' which mbasic uses for next pages
            more_btn = driver.find_element(By.XPATH, "//a[contains(@href, 'pager')]")
            more_btn.click()
            time.sleep(2)
        except:
            break

    # 2. Click all 'See more' links within comments to get full text
    see_mores = driver.find_elements(By.LINK_TEXT, "See more")
    for link in see_mores:
        try:
            driver.execute_script("arguments[0].click();", link)
        except:
            pass


# ================= STRUCTURED EXTRACTION =================
def extract_structured_comments(driver):
    """Targets individual comment containers to separate Author from Text."""
    # Foundational logic: Each comment on mbasic is a div containing an h3 (Author)
    comment_blocks = driver.find_elements(By.XPATH, "//div[h3 and div]")
    
    results = []
    for block in comment_blocks:
        try:
            # Author is the text inside the h3 tag
            author = block.find_element(By.TAG_NAME, "h3").text.strip()
            
            # Content is the div that isn't the footer actions (Like/Reply)
            # Usually the first child div after the h3
            content_el = block.find_element(By.XPATH, "./div[1]")
            content = content_el.text.strip()
            
            # Detect Reply based on indentation style
            style = block.get_attribute("style") or ""
            ctype = "REPLY" if "margin-left" in style.lower() or "padding-left" in style.lower() else "COMMENT"

            # Filter out UI system messages (e.g., 'View 4 more comments')
            if author and content and "View all" not in author:
                results.append((author, content, ctype))
        except:
            continue
    return results


# ================= MAIN LOOP =================
def run():
    wb_in = load_workbook(INPUT_EXCEL)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append(["Source", "Keyword", "Commentator", "Comment", "Type", "Post URL"])

    driver = init_driver()
    load_cookies(driver)

    try:
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            post_url = row[1]
            if not post_url: continue

            mbasic_url = post_url.replace("www.facebook.com", "mbasic.facebook.com")
            print(f"Processing: {mbasic_url}")

            driver.get(mbasic_url)
            time.sleep(3)

            expand_everything(driver)
            data = extract_structured_comments(driver)

            for author, comment, ctype in data:
                ws_out.append([SOURCE, KEYWORD, author, comment, ctype, post_url])

    finally:
        wb_out.save(OUTPUT_EXCEL)
        driver.quit()
        print(f"DONE. Results saved to: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    run()
