import os
import time
from glob import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
OUTPUT_DIR = "output"
FINAL_EXCEL = "fb_comments_final.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

MAX_POSTS = 20
SOURCE = "Facebook"
KEYWORD = "probate"


# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )


# ================= LOAD COOKIES =================
def load_driver_with_cookies():
    driver = init_driver()
    driver.get("https://www.facebook.com/")
    time.sleep(4)

    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    parts = line.strip().split("\t")
                    if len(parts) >= 7:
                        driver.add_cookie({
                            "name": parts[5],
                            "value": parts[6],
                            "domain": parts[0]
                        })
        driver.refresh()
        time.sleep(6)
        print("Cookies loaded successfully")
    else:
        print("Cookie file not found")

    return driver


# ================= FIND POST URL EXCEL =================
def get_post_excel():
    files = glob(os.path.join(OUTPUT_DIR, "*.xlsx"))
    for f in sorted(files, key=os.path.getmtime, reverse=True):
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        if "Post URL" in headers:
            print(f"Using input Excel: {f}")
            return f
    raise Exception("No Excel with 'Post URL' column found")


# ================= READ POST URLS =================
def read_post_urls():
    excel = get_post_excel()
    wb = load_workbook(excel)
    ws = wb.active

    post_col = [c.value for c in ws[1]].index("Post URL")

    urls = []
    for r in ws.iter_rows(min_row=2):
        if len(urls) >= MAX_POSTS:
            break
        if r[post_col].value:
            urls.append(r[post_col].value)

    print(f"Loaded {len(urls)} post URLs")
    return urls


# ================= CLICK VIEW MORE =================
def click_all_view_more(driver, rounds=15):
    for _ in range(rounds):
        buttons = driver.find_elements(
            By.XPATH,
            "//span[contains(text(),'View') and (contains(text(),'comment') or contains(text(),'repl'))]"
        )
        clicked = 0
        for btn in buttons:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(1)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(2)
                clicked += 1
            except Exception:
                continue
        if clicked == 0:
            break


# ================= EXTRACT COMMENTS =================
def extract_comments(driver, post_url, ws):
    print(f"Opening post: {post_url}")
    driver.get(post_url)
    time.sleep(8)

    # Wait for comment containers (photo-post fix)
    try:
        WebDriverWait(driver, 12).until(
            lambda d: len(d.find_elements(
                By.XPATH,
                "//div[@role='article'] | //div[@data-ad-preview='message']"
            )) > 0
        )
    except Exception:
        pass

    # Slow scroll
    for _ in range(5):
        driver.execute_script("window.scrollBy(0, 400);")
        time.sleep(2)

    click_all_view_more(driver)

    blocks = driver.find_elements(
        By.XPATH,
        "//div[@role='article'] | //div[@data-ad-preview='message']"
    )

    print(f"Comments found: {len(blocks)}")

    for block in blocks:
        try:
            profile = block.find_element(By.XPATH, ".//a[@role='link']")
            commenter_name = profile.text.strip()

            spans = block.find_elements(By.XPATH, ".//span[contains(@class,'x1lliihq')]")
            comment_text = ""

            for sp in spans:
                txt = sp.text.strip()
                if txt and txt != commenter_name:
                    comment_text = txt
                    break

            if not commenter_name or not comment_text:
                continue

            ws.append([
                SOURCE,
                KEYWORD,
                "",
                commenter_name,
                post_url,
                comment_text,
                "",
                "",
                "",
                "",
                "",
            ])

        except Exception:
            continue


# ================= MAIN =================
def run():
    driver = load_driver_with_cookies()
    post_urls = read_post_urls()

    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    headers = [
        "Source",
        "Keyword",
        "Group",
        "Commenter Name",
        "Url to find the Comment",
        "Comment",
        "CHATGPT response for Comment",
        "First Name of the Responder1",
        "Last Name of the Responder1",
        "Comment",
        "CHATGPT response for Comments Reponse",
    ]

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for url in post_urls:
        extract_comments(driver, url, ws)

    wb.save(FINAL_EXCEL)
    print(f"EXCEL SAVED: {FINAL_EXCEL}")

    driver.quit()


if __name__ == "__main__":
    run()
