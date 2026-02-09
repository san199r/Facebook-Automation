import os
import time
import re
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
SOURCE = "FB"
KEYWORD = "PROBATE"

INPUT_EXCEL = "output/fb_probate_ALL_posts_20260209_110948.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
SCREENSHOT_DIR = os.path.join(OUTPUT_DIR, "screenshots")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR, f"fb_BATCH_post_comments_{TIMESTAMP}.xlsx"
)


# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-notifications")
    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )


def load_cookies(driver):
    driver.get("https://mbasic.facebook.com/")
    time.sleep(4)

    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.strip() and not line.startswith("#"):
                    parts = line.strip().split("\t")
                    if len(parts) >= 7:
                        driver.add_cookie({
                            "name": parts[5],
                            "value": parts[6],
                            "domain": ".facebook.com"
                        })

    driver.refresh()
    time.sleep(5)


def to_mbasic(url):
    return url.replace("www.facebook.com", "mbasic.facebook.com")


# ================= LOAD ALL COMMENTS (WITH SCREENSHOTS) =================
def load_all_comments(driver, post_idx):
    step = 1
    driver.save_screenshot(
        os.path.join(SCREENSHOT_DIR, f"post_{post_idx:02d}_{step:02d}_open.png")
    )
    step += 1

    for round_no in range(1, 30):
        clicked_any = False

        links = driver.find_elements(By.XPATH, "//a[contains(text(),'View more comments')]")
        for link in links:
            try:
                link.click()
                clicked_any = True
                time.sleep(3)
                driver.save_screenshot(
                    os.path.join(SCREENSHOT_DIR, f"post_{post_idx:02d}_{step:02d}_view_more.png")
                )
                step += 1
            except:
                continue

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)
        driver.save_screenshot(
            os.path.join(SCREENSHOT_DIR, f"post_{post_idx:02d}_{step:02d}_scroll.png")
        )
        step += 1

        if not clicked_any:
            break


# ================= CLEAN PARSER =================
def parse_comments_from_body(body_text):
    lines = [l.strip() for l in body_text.splitlines() if l.strip()]

    ignore = {
        "like", "reply", "share", "comment", "most relevant",
        "view more comments", "write a comment", "comments"
    }

    cleaned = []
    for line in lines:
        low = line.lower()
        if low in ignore:
            continue
        if re.match(r"\d+w", low):
            continue
        if re.match(r"\d+\s+of\s+\d+", low):
            continue
        if "write a comment" in low:
            continue
        if line.strip() in {"·", ".", "…"}:
            continue
        if line.isdigit():
            continue
        cleaned.append(line)

    pairs = []
    i = 0
    while i < len(cleaned) - 1:
        name = cleaned[i]
        comment = cleaned[i + 1]
        if len(name.split()) <= 4 and len(comment.split()) > 2:
            pairs.append((name, comment))
            i += 2
        else:
            i += 1

    return pairs


# ================= MAIN =================
def run():
    wb_in = load_workbook(INPUT_EXCEL)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Comments"

    headers = [
        "Source",
        "Keyword",
        "Commentator",
        "Comment",
        "Commenter URL",
        "Post URL"
    ]
    ws_out.append(headers)
    for c in ws_out[1]:
        c.font = Font(bold=True)

    driver = init_driver()
    load_cookies(driver)

    try:
        for idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), 1):
            post_url = row[1]
            mbasic_url = to_mbasic(post_url)

            print(f"[{idx}] Opening:", mbasic_url)
            driver.get(mbasic_url)
            time.sleep(6)

            load_all_comments(driver, idx)

            body_text = driver.find_element("tag name", "body").text
            pairs = parse_comments_from_body(body_text)

            for name, comment in pairs:
                ws_out.append([
                    SOURCE,
                    KEYWORD,
                    name,
                    comment,
                    post_url,
                    post_url
                ])

    finally:
        wb_out.save(OUTPUT_EXCEL)
        driver.quit()

        print("===================================")
        print("DONE – BATCH COMMENTS EXTRACTED")
        print("Excel:", OUTPUT_EXCEL)
        print("Screenshots:", SCREENSHOT_DIR)
        print("===================================")


if __name__ == "__main__":
    run()
