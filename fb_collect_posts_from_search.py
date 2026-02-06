import os
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
KEYWORD = "probate"
SEARCH_URL = f"https://www.facebook.com/search/posts/?q={KEYWORD}"

COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
SCREENSHOT_DIR = os.path.join(OUTPUT_DIR, "screenshots")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR, f"facebook_posts_{KEYWORD}_{TIMESTAMP}.xlsx"
)


# ================= DRIVER =================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.set_page_load_timeout(60)
    return driver


# ================= COOKIES =================
def load_facebook_cookies(driver):
    driver.get("https://www.facebook.com/")
    time.sleep(5)

    with open(COOKIE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue

            domain, flag, path, secure, expiry, name, value = line.strip().split("\t")

            cookie = {
                "name": name,
                "value": value,
                "domain": domain,
                "path": path
            }

            if expiry.isdigit():
                cookie["expiry"] = int(expiry)

            try:
                driver.add_cookie(cookie)
            except Exception:
                pass

    driver.refresh()
    time.sleep(8)


# ================= EXCEL =================
def init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Post URLs"

    bold = Font(bold=True)
    ws.append(["S.No", "Post URL"])
    ws["A1"].font = bold
    ws["B1"].font = bold

    return wb, ws


# ================= POST EXTRACTION (FIXED) =================
def collect_post_urls(driver, scrolls=12):
    post_urls = set()

    for i in range(scrolls):
        print(f"Scrolling {i + 1}/{scrolls}")

        articles = driver.find_elements(By.XPATH, "//div[@role='article']")

        for article in articles:
            try:
                links = article.find_elements(By.XPATH, ".//a[@href]")
                for a in links:
                    href = a.get_attribute("href")
                    if not href:
                        continue

                    clean = href.split("?")[0]

                    if (
                        "facebook.com" in clean
                        and "/search/" not in clean
                        and (
                            "/posts/" in clean
                            or "permalink.php" in clean
                            or "story_fbid=" in clean
                        )
                    ):
                        post_urls.add(clean)
                        break  # one post URL per article
            except Exception:
                continue

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)

    return post_urls


# ================= MAIN =================
def run():
    driver = init_driver()
    load_facebook_cookies(driver)

    print("Opening Facebook search...")
    driver.get(SEARCH_URL)
    time.sleep(10)

    driver.save_screenshot(
        os.path.join(SCREENSHOT_DIR, f"after_search_{TIMESTAMP}.png")
    )

    print("Collecting post URLs...")
    post_urls = collect_post_urls(driver)

    wb, ws = init_excel()
    for idx, url in enumerate(sorted(post_urls), start=1):
        ws.append([idx, url])

    wb.save(OUTPUT_EXCEL)

    print(f"Total posts collected: {len(post_urls)}")
    print(f"Excel saved: {OUTPUT_EXCEL}")

    driver.save_screenshot(
        os.path.join(SCREENSHOT_DIR, f"before_close_{TIMESTAMP}.png")
    )

    driver.quit()


# ================= RUN =================
if __name__ == "__main__":
    run()
