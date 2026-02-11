import os
import time
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ================= CONFIG =================
INPUT_EXCEL = "clean_posts.xlsx"
OUTPUT_EXCEL = "fb_comments_final.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

SOURCE = "Facebook"
KEYWORD = "probate"

# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-blink-features=AutomationControlled")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

# ================= LOAD COOKIES =================
def load_driver_with_cookies():
    driver = init_driver()
    driver.get("https://mbasic.facebook.com/")
    time.sleep(3)

    if os.path.exists(COOKIE_FILE):
        with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if line.startswith("#") or not line.strip():
                    continue
                parts = line.strip().split("\t")
                if len(parts) >= 7:
                    try:
                        driver.add_cookie({
                            "name": parts[5],
                            "value": parts[6],
                            "domain": ".facebook.com",
                            "path": "/"
                        })
                    except:
                        pass

    driver.refresh()
    time.sleep(5)
    print("Cookies loaded successfully")
    return driver

# ================= READ URLS =================
def read_urls():
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            urls.append(row[0])

    print("Total Posts:", len(urls))
    return urls

# ================= LOAD ALL COMMENTS =================
def load_all_comments(driver):
    while True:
        try:
            more = driver.find_element(By.XPATH, "//a[contains(text(),'View previous comments')]")
            more.click()
            time.sleep(3)
        except:
            break

# ================= EXTRACT COMMENTS =================
def extract_comments(driver, post_url, ws):

    # Convert to mbasic
    mbasic_url = post_url.replace("www.facebook.com", "mbasic.facebook.com")
    driver.get(mbasic_url)
    time.sleep(5)

    load_all_comments(driver)

    comment_blocks = driver.find_elements(By.XPATH, "//div[contains(@id,'comment')]")

    print("Comments found:", len(comment_blocks))

    if not comment_blocks:
        ws.append([SOURCE, KEYWORD, "", "NO_COMMENTS", post_url, "NO_COMMENTS"])
        return

    for block in comment_blocks:
        try:
            name = block.find_element(By.XPATH, ".//h3").text.strip()
            full_text = block.text.strip()

            # Remove name from comment text
            comment = full_text.replace(name, "").strip()

            ws.append([SOURCE, KEYWORD, "", name, post_url, comment])

        except:
            continue

# ================= MAIN =================
def run():
    driver = load_driver_with_cookies()
    urls = read_urls()

    wb = Workbook()
    ws = wb.active
    ws.append(["Source", "Keyword", "Group", "Commenter Name", "Post URL", "Comment"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i, url in enumerate(urls, 1):
        print(f"\nProcessing {i}/{len(urls)}")
        extract_comments(driver, url, ws)
        time.sleep(5)

    wb.save(OUTPUT_EXCEL)
    print("\nSaved:", OUTPUT_EXCEL)
    driver.quit()

if __name__ == "__main__":
    run()
