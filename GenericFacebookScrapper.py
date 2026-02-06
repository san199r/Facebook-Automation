# save as facebook_followers_scraper.py

import os
import re
import time
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
START_URL = "https://www.facebook.com/UseApolloIo/followers/"  # CHANGE ONLY THIS
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def get_safe_filename(url):
    parsed = urlparse(url)
    name = parsed.path.strip("/").replace("/", "_")
    return name if name else "facebook_followers"


EXCEL_FILE = os.path.join(
    OUTPUT_DIR,
    f"{get_safe_filename(START_URL)}.xlsx"
)

HEADERS = [
    "S.No",
    "Facebook Name",
    "Facebook Profile URL",
]


# ================= UTILS =================
def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="ignore").decode())


def normalize(text):
    return re.sub(r"\s+", " ", (text or "").strip())


def is_valid_name(name):
    bad = {
        "followers", "following", "about", "mentions",
        "reviews", "reels", "photos", "home"
    }
    return name and name.lower() not in bad and len(name) > 2


# ================= DRIVER =================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    driver.set_page_load_timeout(60)
    return driver


# ================= COOKIES =================
def load_facebook_cookies(driver):
    driver.get("https://www.facebook.com/")
    time.sleep(3)

    with open(COOKIE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue

            domain, flag, path, secure, expiry, name, value = line.strip().split("\t")

            if not domain.endswith("facebook.com"):
                continue

            cookie = {
                "name": name,
                "value": value,
                "domain": domain,
                "path": path,
            }

            if expiry.isdigit():
                cookie["expiry"] = int(expiry)

            driver.add_cookie(cookie)

    driver.refresh()
    time.sleep(5)


# ================= EXCEL =================
def init_or_resume_excel():
    collected = set()

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        sno = ws.max_row

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[2]:
                collected.add(row[2])

        safe_print(f"Resuming with {len(collected)} followers")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Followers"

        bold = Font(bold=True)
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(1, i, h).font = bold

        sno = 1
        safe_print("Creating new followers Excel")

    return wb, ws, collected, sno


# ================= MAIN =================
def scrape_followers():
    driver = init_driver()
    load_facebook_cookies(driver)

    if "login" in driver.current_url.lower():
        raise Exception("❌ Cookie login failed")

    driver.get(START_URL)
    time.sleep(5)

    wb, ws, collected, sno = init_or_resume_excel()

    no_new_rounds = 0
    MAX_NO_NEW = 15
    last_height = 0

    safe_print("Collecting followers...")

    while no_new_rounds < MAX_NO_NEW:
        found = 0

        anchors = driver.find_elements(
            By.XPATH,
            "//a[contains(@href,'facebook.com') and "
            "(contains(@href,'/profile.php') or contains(@href,'/people/'))]"
        )

        for a in anchors:
            try:
                name = normalize(a.text)
                href = a.get_attribute("href")

                if not is_valid_name(name):
                    continue
                if not href or href in collected:
                    continue

                collected.add(href)
                ws.append([sno, name, href])
                safe_print(f"Collected {sno}: {name}")

                sno += 1
                found += 1

            except StaleElementReferenceException:
                continue

        if found == 0:
            no_new_rounds += 1
        else:
            no_new_rounds = 0

        driver.execute_script("window.scrollBy(0, 800);")
        time.sleep(2)

        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            no_new_rounds += 1
        last_height = new_height

    wb.save(EXCEL_FILE)
    driver.quit()
    safe_print(f"✅ Followers scraping completed: {EXCEL_FILE}")


if __name__ == "__main__":
    scrape_followers()
