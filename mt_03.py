import os
import time
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
COOKIE_FILE = "cookies/facebook_cookies.json"
POSTS_FILE = "facebook_posts_input.xlsx"   # Excel with Post URLs in column C
OUTPUT_FILE = "fb_comments_structured.xlsx"


# ================= DRIVER =================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--headless=new")  # Required for Jenkins

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )


# ================= COOKIE LOGIN =================
def load_cookies(driver):
    print("Opening Facebook homepage...")
    driver.get("https://www.facebook.com/")
    time.sleep(3)

    print("Loading cookies...")
    with open(COOKIE_FILE, "r", encoding="utf-8") as f:
        cookies = json.load(f)

    for cookie in cookies:
        cookie.pop("sameSite", None)
        try:
            driver.add_cookie(cookie)
        except Exception:
            continue

    driver.refresh()
    time.sleep(5)

    print("Current URL after cookie load:", driver.current_url)


# ================= LOAD POST URLS =================
def load_post_urls():
    from openpyxl import load_workbook

    wb = load_workbook(POSTS_FILE)
    ws = wb.active

    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            urls.append(row[2])

    return urls


# ================= SCROLL COMMENTS =================
def scroll_comments(driver):
    last_height = driver.execute_script("return document.body.scrollHeight")

    for _ in range(8):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


# ================= EXPAND SEE MORE =================
def expand_see_more(driver):
    buttons = driver.find_elements(
        By.XPATH,
        "//div[@role='button']//span[contains(text(),'See more')]"
    )

    for b in buttons:
        try:
            driver.execute_script("arguments[0].click();", b)
            time.sleep(0.3)
        except Exception:
            continue


# ================= EXTRACT COMMENTS =================
def extract_comments(driver):
    comments = set()

    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "//div[@role='article']"))
        )
    except Exception:
        return comments

    scroll_comments(driver)
    expand_see_more(driver)

    comment_elements = driver.find_elements(
        By.XPATH,
        "//div[@role='article']//div[@dir='auto']"
    )

    for c in comment_elements:
        try:
            text = c.text.strip()
            if len(text) > 5:
                comments.add(text)
        except Exception:
            continue

    return comments


# ================= INIT OUTPUT FILE =================
def init_output():
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    headers = ["Post URL", "Comment"]

    bold = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header).font = bold
        ws.cell(1, col, header).value = header

    wb.save(OUTPUT_FILE)
    return wb, ws


# ================= MAIN =================
def main():
    driver = None

    try:
        wb, ws = init_output()
        driver = init_driver()

        load_cookies(driver)

        post_urls = load_post_urls()
        print("Total Posts:", len(post_urls))

        total_comments = 0

        for i, url in enumerate(post_urls, start=1):
            print(f"Processing {i}/{len(post_urls)}")
            print("Opening:", url)

            driver.get(url)
            time.sleep(4)

            if "login" in driver.current_url:
                print("Not logged in. Skipping post.")
                continue

            comments = extract_comments(driver)
            print("Extracted:", len(comments))

            for comment in comments:
                ws.append([url, comment])
                total_comments += 1

            wb.save(OUTPUT_FILE)

        print("Saved:", OUTPUT_FILE)
        print("Total Unique Comments:", total_comments)

    finally:
        if driver:
            driver.quit()


if __name__ == "__main__":
    main()
