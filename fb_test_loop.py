import os
import time
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# ================= CONFIG =================
KEYWORD = "probate"
SOURCE = "Facebook"

INPUT_EXCEL = "facebook_fixed.xlsx"
FINAL_EXCEL = "fb_comments_final.xlsx"
COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

MAX_POSTS = 20
SCROLL_COUNT = 5

# ================= DRIVER =================
def init_driver():
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--window-size=1200,900")
    options.add_argument("--disable-blink-features=AutomationControlled")

    return webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )

# ================= LOAD COOKIES =================
def load_driver_with_cookies():
    driver = init_driver()
    driver.get("https://mbasic.facebook.com/")
    time.sleep(3)

    if not os.path.exists(COOKIE_FILE):
        print("Cookie file not found!")
        return driver

    with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue
            parts = line.strip().split("\t")
            if len(parts) >= 7:
                try:
                    driver.add_cookie({
                        "name": parts[5],
                        "value": parts[6],
                        "domain": ".facebook.com",
                        "path": "/"
                    })
                except:
                    continue

    driver.refresh()
    time.sleep(5)
    print("Cookies loaded successfully")
    return driver

# ================= READ POST URLS =================
def read_post_urls():
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    if "Post URL" not in headers:
        raise Exception("Post URL column not found!")

    col = headers.index("Post URL") + 1

    urls = []
    for row in ws.iter_rows(min_row=2):
        if len(urls) >= MAX_POSTS:
            break
        val = row[col - 1].value
        if val:
            urls.append(str(val).strip())

    print(f"Total URLs Loaded: {len(urls)}")
    return urls

# ================= SCROLL =================
def scroll_comments(driver):
    for _ in range(SCROLL_COUNT):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)

# ================= CLICK MORE =================
def expand_more_comments(driver):
    try:
        links = driver.find_elements(By.XPATH, "//a[contains(text(),'View more comments')]")
        for link in links:
            try:
                link.click()
                time.sleep(3)
            except:
                continue
    except:
        pass

# ================= FBID EXTRACTION =================
def extract_fbid(post_url):
    match = re.search(r"fbid=(\d+)", post_url)
    if match:
        return match.group(1)
    return None

# ================= EXTRACT COMMENTS =================
def extract_comments(driver, post_url, ws):
    fbid = extract_fbid(post_url)
    if not fbid:
        print("FBID not found:", post_url)
        return

    mbasic_url = f"https://mbasic.facebook.com/photo.php?fbid={fbid}"
    print("Opening:", mbasic_url)

    driver.get(mbasic_url)
    time.sleep(5)

    scroll_comments(driver)
    expand_more_comments(driver)

    comment_blocks = driver.find_elements(By.XPATH, "//div[contains(@id,'comment')]")

    if not comment_blocks:
        ws.append([
            SOURCE, KEYWORD, "",
            "NO_COMMENTS",
            post_url,
            "NO_COMMENTS",
            "", "", "", "", ""
        ])
        return

    for block in comment_blocks:
        try:
            name = block.find_element(By.XPATH, ".//h3").text.strip()
            comment = block.find_element(By.XPATH, ".//div[contains(@id,'comment')]").text.strip()

            ws.append([
                SOURCE,
                KEYWORD,
                "",
                name,
                post_url,
                comment,
                "",
                "",
                "",
                "",
                ""
            ])
        except:
            continue

# ================= MAIN =================
def run():
    driver = load_driver_with_cookies()
    post_urls = read_post_urls()

    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"

    headers = [
        "Source", "Keyword", "Group",
        "Commenter Name", "Url to find the Comment",
        "Comment", "CHATGPT response for Comment",
        "First Name of the Responder1",
        "Last Name of the Responder1",
        "Comment",
        "CHATGPT response for Comments Reponse"
    ]

    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for i, url in enumerate(post_urls, start=1):
        print(f"Processing {i}/{len(post_urls)}")
        extract_comments(driver, url, ws)
        time.sleep(5)

    wb.save(FINAL_EXCEL)
    print("EXCEL SAVED:", FINAL_EXCEL)
    driver.quit()

if __name__ == "__main__":
    run()
