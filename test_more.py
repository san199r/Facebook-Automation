import os
import time
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ================= CONFIG =================
SOURCE = "FB"
KEYWORD = "PROBATE"

INPUT_EXCEL = "fb_probate_ALL_posts_20260210_115220.xlsx"
POST_URL_COLUMN = "Post URL"

COOKIE_FILE = os.path.join("cookies", "facebook_cookies.txt")

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL = os.path.join(
    OUTPUT_DIR, f"fb_ALL_posts_comments_{TIMESTAMP}.xlsx"
)


# ================= DRIVER (DESKTOP) =================
def init_driver():
    options = Options()
    options.add_argument("--window-size=1920,1080")          # full desktop resolution
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    # Hide webdriver property so FB doesn't detect automation
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


# ================= COOKIES =================
def load_cookies(driver):
    driver.get("https://www.facebook.com/")
    time.sleep(3)

    if not os.path.exists(COOKIE_FILE):
        print("[WARN] No cookie file found at:", COOKIE_FILE)
        return

    with open(COOKIE_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            if line.strip() and not line.startswith("#"):
                p = line.strip().split("\t")
                if len(p) >= 7:
                    try:
                        driver.add_cookie({
                            "name": p[5],
                            "value": p[6],
                            "domain": ".facebook.com"
                        })
                    except Exception:
                        pass

    driver.refresh()
    time.sleep(4)


# ================= URL HELPER =================
def to_desktop_url(url):
    """Ensure URL uses www.facebook.com (not mbasic)."""
    return url.replace("mbasic.facebook.com", "www.facebook.com")


# ================= HELPERS =================
def is_timestamp(text):
    t = text.lower().strip()
    if re.match(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{1,2},?\s*\d{4}", t
    ):
        return True
    if re.match(
        r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{1,2}$", t
    ):
        return True
    if re.match(r"^\d+\s*(h|hr|hrs|min|mins|d|w|m|y|s)\b", t):
        return True
    if t in ("just now", "yesterday", "today"):
        return True
    return False


def is_ui_noise(text):
    t = text.strip()
    if not t:
        return True
    if re.match(r"^[\W_]+$", t):
        return True
    if re.match(r"^\d+$", t):
        return True
    ui_words = {
        "like", "reply", "share", "watch", "follow", "send", "more",
        "see more", "view more", "load more", "write a comment", "comment",
        "reactions", "top comments", "most relevant", "newest", "all comments",
        "hide", "report", "unlike", "love", "haha", "wow", "sad", "angry",
        "view", "comments", "comment on this", "add a comment",
        "view all comments", "view more comments", "load more comments",
    }
    if t.lower() in ui_words:
        return True
    if re.match(r"^[\uE000-\uF8FF\U0001F300-\U0001FAFF\s]+$", t):
        return True
    return False


def has_real_text(text):
    return bool(re.search(r"[A-Za-z]", text))


def clean_comment_text(raw):
    lines = raw.splitlines()
    cleaned = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if is_ui_noise(line) or is_timestamp(line):
            continue
        line = re.sub(r"[\uE000-\uF8FF\U0001F300-\U0001FAFF]\s*\d*\s*", "", line).strip()
        if line:
            cleaned.append(line)
    return " ".join(cleaned).strip()


# ================= SCROLL & EXPAND COMMENTS =================
def expand_all_comments(driver):
    """
    On desktop FB, comments may be hidden behind:
    - 'View X more comments' buttons
    - 'See more' links inside comment text
    This function keeps clicking them until none remain.
    """
    wait = WebDriverWait(driver, 5)
    max_attempts = 30
    attempt = 0

    while attempt < max_attempts:
        attempt += 1
        clicked = False

        # Click "View X more comments" / "View more comments"
        try:
            btns = driver.find_elements(
                By.XPATH,
                "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                "'abcdefghijklmnopqrstuvwxyz'),'view') and "
                "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                "'abcdefghijklmnopqrstuvwxyz'),'comment')]"
            )
            for btn in btns:
                try:
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(2)
                    clicked = True
                except:
                    pass
        except:
            pass

        # Click "Load more comments"
        try:
            btns = driver.find_elements(
                By.XPATH,
                "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ',"
                "'abcdefghijklmnopqrstuvwxyz'),'load more')]"
            )
            for btn in btns:
                try:
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(2)
                    clicked = True
                except:
                    pass
        except:
            pass

        if not clicked:
            break

    # Scroll to bottom to trigger lazy-loaded comments
    last_height = driver.execute_script("return document.body.scrollHeight")
    for _ in range(5):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


# ================= DESKTOP DOM COMMENT EXTRACTOR =================
def extract_comments_desktop(driver):
    """
    Extract comments from desktop Facebook.

    Desktop FB renders comments inside aria-label elements.
    Key selectors:
    - div[aria-label='Comment by ...'] — each comment block
    - Commenter name: first <a> or <span> with the name
    - Comment text: deeper span with actual text content
    """
    results = []

    # --- Strategy 1: aria-label="Comment by NAME" ---
    try:
        comment_blocks = driver.find_elements(
            By.XPATH, "//div[starts-with(@aria-label,'Comment by')]"
        )
        if comment_blocks:
            for block in comment_blocks:
                try:
                    aria = block.get_attribute("aria-label") or ""
                    # Extract name from aria-label: "Comment by John Smith"
                    name_match = re.match(r"Comment by (.+)", aria)
                    name = name_match.group(1).strip() if name_match else ""

                    if not name:
                        # Fallback: get name from first link in block
                        links = block.find_elements(By.XPATH, ".//a[@role='link']")
                        if links:
                            name = links[0].text.strip()

                    if not name or not has_real_text(name):
                        continue

                    # Get comment text — look for the deepest text span
                    text_spans = block.find_elements(
                        By.XPATH, ".//div[@dir='auto'] | .//span[@dir='auto']"
                    )
                    comment_text = ""
                    for span in text_spans:
                        raw = span.text.strip()
                        if not raw or raw == name:
                            continue
                        cleaned = clean_comment_text(raw)
                        if cleaned and has_real_text(cleaned) and len(cleaned) > 2:
                            comment_text = cleaned
                            break

                    if name and comment_text:
                        results.append((name, comment_text))

                except:
                    continue

            if results:
                return results
    except Exception as e:
        print(f"  [WARN] Strategy 1 error: {e}")

    # --- Strategy 2: data-testid or role="article" comment containers ---
    try:
        comment_blocks = driver.find_elements(
            By.XPATH,
            "//div[@role='article'] | //li[.//a[contains(@href,'/user/') or "
            "contains(@href,'profile.php')]]"
        )
        seen = set()
        for block in comment_blocks:
            try:
                # Get commenter name from profile link
                links = block.find_elements(
                    By.XPATH,
                    ".//a[contains(@href,'facebook.com') or "
                    "contains(@href,'/user/') or contains(@href,'profile')]"
                )
                name = ""
                for link in links:
                    t = link.text.strip()
                    if t and has_real_text(t) and len(t.split()) <= 6:
                        name = t
                        break

                if not name or name in seen:
                    continue

                # Get comment text
                text_els = block.find_elements(
                    By.XPATH, ".//div[@dir='auto'] | .//span[@dir='auto']"
                )
                comment_text = ""
                for el in text_els:
                    raw = el.text.strip()
                    if not raw or raw == name:
                        continue
                    cleaned = clean_comment_text(raw)
                    if cleaned and has_real_text(cleaned) and len(cleaned) > 2:
                        comment_text = cleaned
                        break

                if name and comment_text:
                    seen.add(name)
                    results.append((name, comment_text))

            except:
                continue

        if results:
            return results
    except Exception as e:
        print(f"  [WARN] Strategy 2 error: {e}")

    # --- Strategy 3: Generic div[dir='auto'] text scan ---
    try:
        results = generic_text_scan(driver)
    except Exception as e:
        print(f"  [WARN] Strategy 3 error: {e}")

    return results


def generic_text_scan(driver):
    """
    Broad fallback: find all profile links, then grab the adjacent text.
    """
    results = []
    seen = set()

    profile_links = driver.find_elements(
        By.XPATH,
        "//a[contains(@href,'facebook.com/') and not(contains(@href,'photo')) "
        "and not(contains(@href,'video')) and not(contains(@href,'groups'))]"
    )

    for link in profile_links:
        try:
            name = link.text.strip()
            if not name or not has_real_text(name) or name in seen:
                continue
            if len(name.split()) > 6:
                continue

            # Go up to find the comment container
            parent = link
            for _ in range(4):
                parent = parent.find_element(By.XPATH, "..")
                text_els = parent.find_elements(
                    By.XPATH, ".//div[@dir='auto'] | .//span[@dir='auto']"
                )
                comment_text = ""
                for el in text_els:
                    raw = el.text.strip()
                    if not raw or raw == name:
                        continue
                    cleaned = clean_comment_text(raw)
                    if cleaned and has_real_text(cleaned) and len(cleaned) > 5:
                        comment_text = cleaned
                        break
                if comment_text:
                    seen.add(name)
                    results.append((name, comment_text))
                    break

        except:
            continue

    return results


# ================= READ POST URLS =================
def read_post_urls():
    wb = load_workbook(INPUT_EXCEL)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = headers.index(POST_URL_COLUMN)

    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx]:
            urls.append(row[idx])

    return urls


# ================= MAIN =================
def run():
    post_urls = read_post_urls()
    print("Total posts:", len(post_urls))

    driver = init_driver()
    load_cookies(driver)

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Comments"

    headers = ["Source", "Keyword", "Commentator", "Comment", "Post URL"]
    ws_out.append(headers)
    for c in ws_out[1]:
        c.font = Font(bold=True)

    total_comments = 0

    for idx, post_url in enumerate(post_urls, 1):
        print(f"[{idx}/{len(post_urls)}] Processing:", post_url)

        try:
            url = to_desktop_url(post_url)
            driver.get(url)
            time.sleep(4)

            # Expand all comments
            expand_all_comments(driver)

            # Extract
            comments = extract_comments_desktop(driver)

        except Exception as e:
            print(f"  [ERROR] {e}")
            ws_out.append([SOURCE, KEYWORD, "ERROR", str(e), post_url])
            continue

        if not comments:
            ws_out.append([SOURCE, KEYWORD, "NO_COMMENTS", "NO_COMMENTS", post_url])
            print("  -> No comments found")
            continue

        for name, comment in comments:
            ws_out.append([SOURCE, KEYWORD, name, comment, post_url])
            total_comments += 1

        print(f"  -> {len(comments)} comments captured")

    driver.quit()
    wb_out.save(OUTPUT_EXCEL)

    print("===================================")
    print("DONE")
    print("Total comments:", total_comments)
    print("Output Excel:", OUTPUT_EXCEL)
    print("===================================")


if __name__ == "__main__":
    run()
